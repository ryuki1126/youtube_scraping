# 利用するライブラリ
import time
import requests
from bs4 import BeautifulSoup
import openpyxl

# 変数
url_list = []
registrant_list = []
channel_list = []
domain = 'https://ytranking.net'

# スクレイピング処理(ランキング対象チャンネルの抽出)
for url_num in range(30, 31):
  print('ランキング対象チャンネルの抽出を開始')

  url = 'https://ytranking.net/ranking/mon/?p='+str(url_num)+'&mode=view&date=202110' #末尾の年月を書き換える事でどの月のデータを取得するかを決める
  res = requests.get(url)
  soup = BeautifulSoup(res.text, 'lxml')
  ul = soup.find('ul', class_='channel-list')
  
  # ランキングリストのaタグからhrefの情報を抽出する
  for a in ul.find_all('a'):
    rank_link = domain + a.get('href')
    url_list.append(rank_link)
  
  #DOM攻撃判定をされないようアクセス間隔を空ける
  print(str(url_num)+'回目のアクセス終了')
  time.sleep(30)
  print('待機終了')

# 抽出したチャンネルの数を取得
num = len(url_list)
print(num)

# チャンネル情報の取得
print('チャンネル情報の抽出を開始')
for ranker in url_list:
  print('アクセス開始')
  channel = {}
  res = requests.get(ranker)
  soup = BeautifulSoup(res.text, 'lxml')
  span = soup.find('span', class_='subscriber-count').contents[0]
  header = soup.find('header', class_='header')
  h1 = header.find('h1')
  a = h1.find('a')
  channel['name'] = a.contents[0]
  channel['subscriber-count'] = span
  channel['url'] = a.attrs['href']
  channel_list.append(channel)
  #DOM攻撃判定をされないようアクセス間隔を空ける
  print(ranker+'の調査終了')
  time.sleep(30)
  print('待機終了')

# Excelに取得した情報を記述する
wb = openpyxl.load_workbook('scraping.xlsx')
sheet = wb['Sheet1']

for i in range(len(channel_list)):
  sheet.cell(row = i+2, column = 1, value= channel_list[i]['name'])
  sheet.cell(row = i+2, column = 2, value= channel_list[i]['subscriber-count'])
  sheet.cell(row = i+2, column = 3, value= channel_list[i]['url'])

wb.save('scraping.xlsx')

print(channel_list)
